import openpyxl as xls
from openpyxl.chart import BarChart, Reference

wb = xls.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row + 1):
    sheet.cell(row, 5).value = sheet.cell(row, 3).value * 0.9

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=5, max_col=5)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'f2')
wb.save('transaction_copy.xlsx')

